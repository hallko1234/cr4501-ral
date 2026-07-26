#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ral_einlernen.py  --  RAL-Faecher mit dem 3nh CR4501 einlernen

Misst alle 216 RAL-Classic-Toene der Reihe nach, jeden genau dreimal,
und schreibt am Ende eine Excel-Mappe als Grundlage fuer die
geraetegenaue Farbtabelle.

BEDIENUNG
  Geraet per USB anschliessen und kalibrieren.
      python -m pip install pyserial openpyxl
      python ral_einlernen.py COM3

  Das Programm nennt die naechste Farbe. Karte aufsetzen, am Geraet
  messen - mehr nicht. Nach drei sauberen Messungen schaltet es von
  selbst zur naechsten Farbe weiter. Keine Tastatureingabe noetig.

  Abbrechen jederzeit mit Strg+C. Beim naechsten Start geht es an
  derselben Stelle weiter (Zwischenstand in ral_fortschritt.csv).

FIRMWARE GLEICH MITBAUEN
  Gibt man die Originalfirmware des eigenen Geraets mit an, baut das
  Programm nach dem Einlernen sofort die fertige Datei:

      python ral_einlernen.py COM3 --original cr4501_original_....bin

  Die Originaldatei ist die Sicherung, die das Webwerkzeug beim Auslesen
  anlegt (180224 Byte). Das Ergebnis heisst cr4501_ral.bin und laesst sich
  im Webwerkzeug unter "Sicherung einspielen" aufs Geraet bringen.

  Ohne Messen, nur bauen - etwa nach einem Firmware-Update:

      python ral_einlernen.py --nur-bauen --original cr4501_original_....bin

PRUEFUNGEN
  - Wiederholung erkannt: liegt eine Messung sehr nah an der eben
    abgeschlossenen Farbe und deutlich weiter von der erwarteten,
    war die alte Karte noch unter dem Kopf. Wird verworfen.
  - Streuung: die drei Messungen muessen zueinander passen. Reisst
    eine aus, wird sie verworfen und eine Ersatzmessung angefordert.
  - Erwartungsabgleich: passt der Messwert ueberhaupt nicht zur
    erwarteten RAL-Farbe, kommt eine Warnung - meist ist dann die
    Karte verrutscht oder eine Farbe uebersprungen.
"""

import sys, os, csv, math, time, statistics

try:
    import serial
except ImportError:
    serial = None          # erst beim Messen noetig, nicht beim Bauen

# ---------------------------------------------------------------
# Einstellungen
# ---------------------------------------------------------------

BAUD            = 115200
WIEDERHOLUNGEN  = 3       # Messungen je Farbe
DE_STREUUNG     = 1.5     # max. Abstand der Einzelmessungen zueinander
DE_WIEDERHOLUNG = 1.0     # naeher als das an der Vorfarbe = Doppelmessung
DE_WARNUNG      = 18.0    # weiter als das von der Erwartung = Warnung

FORTSCHRITT = "ral_fortschritt.csv"
EXCEL       = "ral_einlernung.xlsx"
RAL_KARTE = [
    (1000, 75.99, -0.63, 27.84, 'CDBA88', 'Gruenbeige'),
    (1001, 73.63, 5.39, 26.86, 'D0B084', 'Beige'),
    (1002, 71.93, 6.94, 36.84, 'D2AA6D', 'Sandgelb'),
    (1003, 75.15, 19.32, 78.76, 'F9A900', 'Signalgelb'),
    (1004, 70.14, 16.15, 74.26, 'E49E00', 'Goldgelb'),
    (1005, 63.62, 13.29, 68.51, 'CB8F00', 'Honiggelb'),
    (1006, 66.48, 22.33, 71.67, 'E19000', 'Maisgelb'),
    (1007, 66.45, 27.41, 72.08, 'E88C00', 'Narzissengelb'),
    (1011, 57.2, 12.53, 32.99, 'AF8050', 'Braunbeige'),
    (1012, 73.67, 4.74, 69.04, 'DDAF28', 'Zitronengelb'),
    (1013, 87.04, 0.42, 10.08, 'E3D9C7', 'Perlweiss'),
    (1014, 80.3, 2.76, 23.84, 'DDC49B', 'Elfenbein'),
    (1015, 85.12, 2.41, 16.9, 'E6D2B5', 'Hellelfenbein'),
    (1016, 87.34, -9.79, 76.74, 'F1DD39', 'Schwefelgelb'),
    (1017, 75.18, 20.28, 55.47, 'F6A951', 'Safrangelb'),
    (1018, 83.33, 3.37, 75.89, 'FACA31', 'Zinkgelb'),
    (1019, 60.73, 4.49, 14.07, 'A48F7A', 'Graubeige'),
    (1020, 59.94, 0.0, 24.69, 'A08F65', 'Olivgelb'),
    (1021, 77.9, 11.23, 80.45, 'F6B600', 'Rapsgelb'),
    (1023, 77.77, 12.18, 80.41, 'F7B500', 'Verkehrsgelb'),
    (1024, 62.17, 8.61, 41.38, 'BA8F4C', 'Ockergelb'),
    (1026, 97.14, -21.56, 94.48, 'FFFF00', 'Leuchtgelb'),
    (1027, 55.51, 6.08, 58.34, 'A77F0F', 'Currygelb'),
    (1028, 72.92, 28.62, 77.65, 'FF9C00', 'Melonengelb'),
    (1032, 71.12, 12.62, 74.81, 'E2A300', 'Ginstergelb'),
    (1033, 71.78, 27.5, 71.82, 'F99A1D', 'Dahliengelb'),
    (1034, 70.9, 22.68, 49.89, 'EB9C52', 'Pastellgelb'),
    (1035, 55.32, 1.28, 11.92, '8F8370', 'Perlbeige'),
    (1036, 44.34, 6.38, 24.65, '806440', 'Perlgold'),
    (1037, 68.81, 27.61, 74.09, 'F09200', 'Sonnengelb'),
    (2000, 58.09, 37.27, 65.95, 'DA6E00', 'Gelborange'),
    (2001, 45.85, 43.97, 47.44, 'BA481C', 'Rotorange'),
    (2002, 44.4, 52.29, 43.64, 'BF3922', 'Blutorange'),
    (2003, 64.31, 43.88, 61.83, 'F67829', 'Pastellorange'),
    (2004, 54.41, 53.1, 63.7, 'E25304', 'Reinorange'),
    (2005, 58.59, 65.1, 67.89, 'FF4D08', 'Leuchtorange'),
    (2007, 77.93, 17.15, 80.94, 'FFB200', 'Leuchthellorange'),
    (2008, 60.31, 46.22, 60.46, 'EC6B22', 'Hellrotorange'),
    (2009, 53.74, 51.73, 62.13, 'DE5308', 'Verkehrsorange'),
    (2010, 53.38, 42.6, 49.82, 'D05D29', 'Signalorange'),
    (2011, 59.3, 40.59, 64.58, 'E26E0F', 'Tieforange'),
    (2012, 56.09, 42.34, 33.83, 'D5654E', 'Lachsorange'),
    (2013, 37.28, 33.84, 32.22, '923E25', 'Perlorange'),
    (2017, 59.15, 61.07, 69.42, 'FC5500', 'RAL orange'),
    (3000, 37.67, 50.31, 36.6, 'A72920', 'Feuerrot'),
    (3001, 34.7, 48.31, 31.03, '9B2423', 'Signalrot'),
    (3002, 34.56, 48.58, 32.06, '9B2321', 'Karminrot'),
    (3003, 29.22, 44.94, 24.29, '861A22', 'Rubinrot'),
    (3004, 23.93, 35.29, 15.9, '6B1C23', 'Purpurrot'),
    (3005, 19.71, 29.78, 12.47, '59191F', 'Weinrot'),
    (3007, 16.37, 14.75, 5.0, '3E2022', 'Schwarzrot'),
    (3009, 29.15, 24.59, 16.13, '6D342D', 'Oxidrot'),
    (3011, 27.78, 36.45, 21.26, '782423', 'Braunrot'),
    (3012, 61.44, 21.76, 22.92, 'C5856D', 'Beigerot'),
    (3013, 35.33, 43.25, 30.32, '972E25', 'Tomatenrot'),
    (3014, 58.27, 34.72, 14.16, 'CB7375', 'Altrosa'),
    (3015, 71.21, 21.64, 5.05, 'D8A0A6', 'Hellrosa'),
    (3016, 40.67, 42.49, 30.66, 'A63D30', 'Korallenrot'),
    (3017, 51.66, 47.2, 19.12, 'CA555D', 'Rose'),
    (3018, 47.05, 54.07, 24.27, 'C63F4A', 'Erdbeerrot'),
    (3020, 40.55, 58.93, 47.95, 'BB1F11', 'Verkehrsrot'),
    (3022, 56.11, 38.53, 29.73, 'CF6955', 'Lachsrot'),
    (3024, 55.33, 74.76, 58.26, 'FF2D21', 'Leuchtrot'),
    (3026, 55.07, 75.33, 60.12, 'FF2A1C', 'Leuchthellrot'),
    (3027, 38.61, 53.51, 21.03, 'AB273C', 'Himbeerrot'),
    (3028, 45.31, 60.8, 44.36, 'CC2C24', 'Reinrot'),
    (3031, 39.31, 46.83, 24.82, 'A63437', 'Orientrot'),
    (3032, 25.11, 36.69, 16.94, '701D24', 'Perlrubinrot'),
    (3033, 39.95, 43.4, 30.96, 'A53A2E', 'Perlrosa'),
    (4001, 45.44, 19.23, -13.91, '816183', 'Rotlila'),
    (4002, 36.75, 35.88, 8.09, '8D3C4B', 'Rotviolett'),
    (4003, 54.22, 44.36, -5.23, 'C4618C', 'Erikaviolett'),
    (4004, 23.61, 34.22, 0.9, '651E38', 'Bordeauxviolett'),
    (4005, 47.09, 16.55, -25.19, '76689A', 'Blaulila'),
    (4006, 36.99, 46.38, -16.81, '903373', 'Verkehrspurpur'),
    (4007, 19.75, 20.47, -8.15, '47243C', 'Purpurviolett'),
    (4008, 40.71, 32.37, -20.5, '844C82', 'Signalviolett'),
    (4009, 58.31, 10.85, -3.18, '9D8692', 'Pastellviolett'),
    (4010, 46.4, 54.19, -4.39, 'BB4077', 'Telemagenta'),
    (4011, 44.22, 12.24, -18.35, '6E6387', 'Perlviolett'),
    (4012, 45.79, 4.07, -11.19, '6A6B7F', 'Perlbrombeer'),
    (5000, 32.58, -1.54, -21.36, '304F6E', 'Violettblau'),
    (5001, 29.87, -9.54, -19.36, '0E4C64', 'Gruenblau'),
    (5002, 24.43, 12.17, -42.0, '00387A', 'Ultramarinblau'),
    (5003, 22.92, 0.47, -20.45, '1F3855', 'Saphirblau'),
    (5004, 11.19, 0.75, -7.51, '191E28', 'Schwarzblau'),
    (5005, 33.85, -0.73, -34.97, '005387', 'Signalblau'),
    (5007, 43.09, -6.94, -23.38, '376B8C', 'Brillantblau'),
    (5008, 23.53, -3.21, -8.15, '2B3A44', 'Graublau'),
    (5009, 37.6, -10.9, -19.83, '215F78', 'Azurblau'),
    (5010, 31.91, -2.75, -31.25, '004F7C', 'Enzianblau'),
    (5011, 16.86, -1.22, -13.01, '1A2B3C', 'Stahlblau'),
    (5012, 53.24, -14.61, -32.24, '0089B6', 'Lichtblau'),
    (5013, 20.16, 3.23, -23.5, '193153', 'Kobaltblau'),
    (5014, 51.29, -3.0, -16.5, '637D96', 'Taubenblau'),
    (5015, 48.9, -9.85, -35.01, '007CAF', 'Himmelblau'),
    (5017, 36.7, -3.77, -33.5, '005B8C', 'Verkehrsblau'),
    (5018, 52.27, -30.1, -9.49, '048B8C', 'Tuerkisblau'),
    (5019, 37.2, -9.27, -27.23, '005E83', 'Capriblau'),
    (5020, 24.6, -14.23, -11.31, '00414B', 'Ozeanblau'),
    (5021, 44.31, -26.37, -9.05, '007577', 'Wasserblau'),
    (5022, 19.86, 10.87, -28.59, '222D5A', 'Nachtblau'),
    (5023, 42.94, -3.48, -23.58, '41698C', 'Fernblau'),
    (5024, 58.35, -10.25, -18.45, '6093AC', 'Pastellblau'),
    (5025, 41.0, -15.42, -17.04, '20697C', 'Perlenzian'),
    (5026, 19.33, 1.72, -24.17, '0F3052', 'Perlnachtblau'),
    (6000, 44.65, -23.68, 5.37, '3C7460', 'Patinagruen'),
    (6001, 39.14, -27.96, 23.11, '366735', 'Smaragdgruen'),
    (6002, 33.93, -24.56, 23.93, '325928', 'Laubgruen'),
    (6003, 34.43, -5.66, 13.11, '50533C', 'Olivgruen'),
    (6004, 25.38, -19.05, -4.2, '024442', 'Blaugruen'),
    (6005, 24.47, -20.84, 5.04, '114232', 'Moosgruen'),
    (6006, 23.95, -0.92, 7.34, '3C392E', 'Grauoliv'),
    (6007, 19.77, -6.12, 9.35, '2C3222', 'Flaschengruen'),
    (6008, 21.61, -1.2, 6.65, '36342A', 'Braungruen'),
    (6009, 20.61, -8.58, 5.06, '27352A', 'Tannengruen'),
    (6010, 43.13, -22.95, 26.14, '4D6F39', 'Grasgruen'),
    (6011, 49.81, -13.03, 17.1, '6B7C59', 'Resedagruen'),
    (6012, 24.47, -6.46, -0.04, '2F3D3A', 'Schwarzgruen'),
    (6013, 49.47, -2.52, 16.29, '7C765A', 'Schilfgruen'),
    (6014, 27.78, 0.25, 8.26, '474135', 'Gelboliv'),
    (6015, 25.55, -1.51, 4.32, '3D3D36', 'Schwarzoliv'),
    (6016, 38.94, -33.87, 9.07, '00694C', 'Tuerkisgruen'),
    (6017, 49.04, -25.79, 29.82, '587F40', 'Maigruen'),
    (6018, 57.64, -35.68, 42.58, '60993B', 'Gelbgruen'),
    (6019, 80.42, -13.27, 14.56, 'B9CEAC', 'Weissgruen'),
    (6020, 26.49, -8.69, 10.12, '37422F', 'Chromoxidgruen'),
    (6021, 61.24, -11.69, 16.12, '8A9977', 'Blassgruen'),
    (6022, 21.63, 0.83, 8.77, '3A3327', 'Braunoliv'),
    (6024, 48.04, -43.17, 18.62, '008351', 'Verkehrsgruen'),
    (6025, 44.01, -15.15, 26.4, '5E6E3B', 'Farngruen'),
    (6026, 35.47, -28.83, 2.91, '005F4E', 'Opalgruen'),
    (6027, 71.48, -20.39, -3.65, '7EBAB5', 'Lichtgruen'),
    (6028, 32.67, -17.34, 6.58, '315442', 'Kieferngruen'),
    (6029, 40.74, -39.84, 20.36, '006F3D', 'Minzgruen'),
    (6032, 47.16, -37.76, 16.95, '237F52', 'Signalgruen'),
    (6033, 51.98, -23.12, -2.21, '45877F', 'Minttuerkis'),
    (6034, 67.28, -16.96, -4.91, '7AADAC', 'Pastelltuerkis'),
    (6035, 28.51, -27.27, 18.69, '194D25', 'Perlgruen'),
    (6036, 32.59, -25.56, 0.62, '04574B', 'Perlopalgruen'),
    (6037, 50.25, -52.43, 41.72, '008B29', 'Reingruen'),
    (6038, 64.26, -65.8, 60.02, '00B51B', 'Leuchtgruen'),
    (6039, 75.71, -23.11, 62.15, 'B3C43E', 'Fasergruen'),
    (7000, 55.79, -3.77, -4.94, '7A888E', 'Fehgrau'),
    (7001, 61.75, -2.89, -3.98, '8C979C', 'Silbergrau'),
    (7002, 50.69, -0.2, 12.8, '817863', 'Olivgrau'),
    (7003, 49.55, -1.3, 7.57, '797669', 'Moosgrau'),
    (7004, 63.9, -0.36, -0.13, '9A9B9B', 'Signalgrau'),
    (7005, 46.1, -1.76, 1.26, '6B6E6B', 'Mausgrau'),
    (7006, 45.59, 2.58, 8.45, '766A5E', 'Beigegrau'),
    (7008, 41.59, 3.57, 22.61, '745F3D', 'Khakigrau'),
    (7009, 40.24, -2.83, 4.16, '5D6058', 'Gruengrau'),
    (7010, 38.55, -2.84, 2.9, '585C56', 'Zeltgrau'),
    (7011, 37.35, -1.83, -3.26, '52595D', 'Eisengrau'),
    (7012, 39.01, -2.12, -1.39, '575D5E', 'Basaltgrau'),
    (7013, 34.36, 0.62, 8.12, '575044', 'Braungrau'),
    (7015, 35.11, -0.46, -3.47, '4F5358', 'Schiefergrau'),
    (7016, 25.8, -1.51, -3.31, '383E42', 'Anthrazitgrau'),
    (7021, 20.57, -0.81, -1.72, '2F3234', 'Schwarzgrau'),
    (7022, 31.47, -0.41, 3.87, '4C4A44', 'Umbragrau'),
    (7023, 53.31, -1.92, 5.42, '808076', 'Betongrau'),
    (7024, 30.83, -0.46, -3.54, '45494E', 'Graphitgrau'),
    (7026, 27.45, -4.34, -2.84, '374345', 'Granitgrau'),
    (7030, 59.11, -0.19, 5.31, '928E85', 'Steingrau'),
    (7031, 43.11, -3.76, -4.48, '5B686D', 'Blaugrau'),
    (7032, 71.86, -0.86, 8.31, 'B5B0A1', 'Kieselgrau'),
    (7033, 53.75, -3.8, 7.19, '7F8274', 'Zementgrau'),
    (7034, 56.95, -0.48, 14.8, '92886F', 'Gelbgrau'),
    (7035, 80.01, -1.23, 1.25, 'C5C7C4', 'Lichtgrau'),
    (7036, 61.24, 1.28, 1.04, '979392', 'Platingrau'),
    (7037, 51.51, -0.57, 0.41, '7A7B7A', 'Staubgrau'),
    (7038, 71.65, -1.29, 3.57, 'B0B0A9', 'Achatgrau'),
    (7039, 43.41, 0.43, 5.22, '6B665E', 'Quarzgrau'),
    (7040, 64.73, -1.54, -2.31, '989EA1', 'Fenstergrau'),
    (7042, 60.21, -1.65, 0.01, '8E9291', 'Verkehrsgrau A'),
    (7043, 34.55, -1.64, 0.75, '4F5250', 'Verkehrsgrau B'),
    (7044, 72.98, -0.55, 6.11, 'B7B3A8', 'Seidengrau'),
    (7045, 60.24, -1.21, -2.22, '8D9295', 'Telegrau 1'),
    (7046, 55.4, -2.09, -3.16, '7E868A', 'Telegrau 2'),
    (7047, 80.58, -0.18, 0.49, 'C8C8C7', 'Telegrau 4'),
    (7048, 51.92, 0.79, 5.18, '817B73', 'Perlmausgrau'),
    (8000, 46.69, 7.4, 28.37, '89693F', 'Gruenbraun'),
    (8001, 47.08, 18.97, 40.11, '9D622B', 'Ockerbraun'),
    (8002, 37.34, 16.75, 16.57, '794D3E', 'Signalbraun'),
    (8003, 37.14, 18.15, 30.1, '7E4B27', 'Lehmbraun'),
    (8004, 38.98, 26.56, 27.08, '8D4931', 'Kupferbraun'),
    (8007, 33.98, 15.03, 23.57, '70462B', 'Rehbraun'),
    (8008, 35.22, 13.06, 28.59, '724A25', 'Olivbraun'),
    (8011, 27.15, 13.04, 16.83, '5A3827', 'Nussbraun'),
    (8012, 27.76, 21.97, 15.39, '66332B', 'Rotbraun'),
    (8014, 24.14, 7.17, 13.13, '4A3526', 'Sepiabraun'),
    (8015, 25.42, 20.36, 15.36, '5E2F26', 'Kastanienbraun'),
    (8016, 21.46, 13.95, 13.64, '4C2B20', 'Mahagonibraun'),
    (8017, 21.66, 8.65, 7.65, '442F29', 'Schokoladenbraun'),
    (8019, 23.3, 2.87, 1.75, '3D3635', 'Graubraun'),
    (8022, 8.15, 1.98, -0.88, '1A1719', 'Schwarzbraun'),
    (8023, 45.45, 28.07, 39.69, 'A45729', 'Orangebraun'),
    (8024, 38.0, 14.31, 21.08, '795038', 'Beigebraun'),
    (8025, 39.94, 9.44, 14.48, '755847', 'Blassbraun'),
    (8028, 26.57, 7.82, 13.92, '513A2A', 'Terrabraun'),
    (8029, 34.8, 25.71, 21.44, '7F4031', 'Perlkupfer'),
    (9001, 89.53, 0.71, 7.93, 'E9E0D2', 'Cremeweiss'),
    (9002, 85.1, -1.42, 5.05, 'D6D5CB', 'Grauweiss'),
    (9003, 93.27, -0.87, 2.41, 'ECECE7', 'Signalweiss'),
    (9004, 17.57, 0.24, -0.65, '2B2B2C', 'Signalschwarz'),
    (9005, 4.02, 0.36, -0.99, '0E0E10', 'Tiefschwarz'),
    (9006, 66.22, -0.19, 0.51, 'A1A1A0', 'Weissaluminium'),
    (9007, 55.51, -0.4, 2.28, '868581', 'Graualuminium'),
    (9010, 93.76, -0.71, 6.27, 'F1EDE1', 'Reinweiss'),
    (9011, 16.46, -0.39, -1.62, '27292B', 'Graphitschwarz'),
    (9012, 95.54, -0.84, 8.9, 'F8F2E1', 'Reinraumweiss'),
    (9016, 94.97, -1.22, 3.36, 'F1F1EA', 'Verkehrsweiss'),
    (9017, 16.62, 0.24, -0.66, '29292A', 'Verkehrsschwarz'),
    (9018, 81.28, -2.3, 3.12, 'C8CBC4', 'Papyrusweiss'),
    (9022, 55.48, -0.39, 1.07, '858583', 'Perlhellgrau'),
    (9023, 51.34, -1.33, 0.15, '787B7A', 'Perldunkelgrau'),
]

# ---------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------

def dE(p, q):
    """Farbabstand zweier Lab-Werte."""
    return math.sqrt((p[0]-q[0])**2 + (p[1]-q[1])**2 + (p[2]-q[2])**2)


def lab_zu_hex(L, a, b):
    """Lab -> sRGB-Hex, nur fuer die Farbvorschau in Excel."""
    y = (L+16)/116; x = a/500 + y; z = y - b/200
    def inv(t): return t**3 if t**3 > 0.008856 else (t-16/116)/7.787
    X = inv(x)*0.95047; Y = inv(y); Z = inv(z)*1.08883
    r =  X*3.2406 - Y*1.5372 - Z*0.4986
    g = -X*0.9689 + Y*1.8758 + Z*0.0415
    bl=  X*0.0557 - Y*0.2040 + Z*1.0570
    def gam(c):
        c = max(0.0, min(1.0, c))
        return 1.055*c**(1/2.4)-0.055 if c > 0.0031308 else 12.92*c
    return "%02X%02X%02X" % (round(gam(r)*255), round(gam(g)*255), round(gam(bl)*255))


def mittel(proben):
    """Mittelwert mehrerer Lab-Messungen."""
    return tuple(statistics.fmean(p[i] for p in proben) for i in range(3))


def bester_treffer(lab):
    """Welche RAL-Nummer der Karte passt am besten zu diesem Messwert?"""
    best = None; bd = 1e9
    for eintrag in RAL_KARTE:
        d = dE(lab, eintrag[1:4])
        if d < bd: bd = d; best = eintrag[0]
    return best, bd


# ---------------------------------------------------------------
# Serielle Schnittstelle
# ---------------------------------------------------------------

import re
MUSTER = re.compile(rb'L\s*=\s*(-?\d+\.?\d*)\s*;\s*a\s*=\s*(-?\d+\.?\d*)\s*;\s*b\s*=\s*(-?\d+\.?\d*)')


class Messgeraet:
    def __init__(self, port):
        if serial is None:
            sys.exit("pyserial fehlt:   python -m pip install pyserial")
        self.ser = serial.Serial(port, BAUD, timeout=0.2)
        self.puffer = b''

    def naechste_messung(self):
        """Blockiert, bis das Geraet eine Messung schickt."""
        while True:
            neu = self.ser.read(512)
            if neu:
                self.puffer += neu
                m = MUSTER.search(self.puffer)
                if m:
                    self.puffer = self.puffer[m.end():]
                    return tuple(float(x) for x in m.groups())

    def leeren(self):
        """Alte Daten verwerfen, damit keine Messung von vorhin zaehlt."""
        time.sleep(0.2)
        self.ser.reset_input_buffer()
        self.puffer = b''

    def schliessen(self):
        self.ser.close()

# ---------------------------------------------------------------
# Fortschritt speichern und laden
# ---------------------------------------------------------------

def fortschritt_laden():
    fertig = {}
    if os.path.exists(FORTSCHRITT):
        with open(FORTSCHRITT, newline='', encoding='utf-8') as f:
            for r in csv.DictReader(f):
                fertig[int(r['ral'])] = {
                    'L': float(r['L']), 'a': float(r['a']), 'b': float(r['b']),
                    'streuung': float(r['streuung']),
                    'abweichung': float(r['abweichung']),
                    'hinweis': r.get('hinweis', ''),
                }
    return fertig


def fortschritt_schreiben(fertig):
    with open(FORTSCHRITT, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['ral', 'L', 'a', 'b', 'streuung', 'abweichung', 'hinweis'])
        for ral in sorted(fertig):
            d = fertig[ral]
            w.writerow([ral, f"{d['L']:.2f}", f"{d['a']:.2f}", f"{d['b']:.2f}",
                        f"{d['streuung']:.2f}", f"{d['abweichung']:.2f}", d['hinweis']])


# ---------------------------------------------------------------
# Eine Farbe einlernen
# ---------------------------------------------------------------

def farbe_einlernen(dev, eintrag, vorfarbe_lab, nummer, gesamt):
    """Sammelt WIEDERHOLUNGEN saubere Messungen. Gibt (mittel, streuung,
    abweichung, hinweis) zurueck - oder None, wenn uebersprungen."""

    ral, eL, ea, eb, ehex, name = eintrag
    erwartet = (eL, ea, eb)

    print(f"\n[{nummer}/{gesamt}]  RAL {ral}  {name}")
    print(f"          Karte aufsetzen und messen ({WIEDERHOLUNGEN}x)")

    proben = []
    verworfen = 0

    while len(proben) < WIEDERHOLUNGEN:
        lab = dev.naechste_messung()

        # --- Wurde versehentlich die vorige Farbe nochmal gemessen? ---
        if vorfarbe_lab is not None and not proben:
            d_alt = dE(lab, vorfarbe_lab)
            d_neu = dE(lab, erwartet)
            if d_alt < DE_WIEDERHOLUNG and d_alt < d_neu:
                print(f"          uebersprungen: das ist noch die vorige Farbe "
                      f"(Abstand {d_alt:.1f}) - bitte die naechste Karte auflegen")
                continue

        # --- Passt der Wert ueberhaupt zur erwarteten Farbe? ---
        d_erw = dE(lab, erwartet)
        if d_erw > DE_WARNUNG and not proben:
            treffer, td = bester_treffer(lab)
            print(f"          Achtung: Messwert liegt {d_erw:.1f} von RAL {ral} entfernt.")
            print(f"          Am besten passt RAL {treffer} (Abstand {td:.1f}).")
            print(f"          Stimmt die Karte? Weitermessen zaehlt den Wert trotzdem.")

        # --- Streuung gegenueber den bisherigen Messungen ---
        if proben:
            d_intern = max(dE(lab, p) for p in proben)
            if d_intern > DE_STREUUNG:
                verworfen += 1
                print(f"          Messung weicht ab ({d_intern:.1f}) - verworfen, "
                      f"bitte wiederholen")
                if verworfen >= 4:
                    print("          Zu viele Ausreisser. Neustart dieser Farbe.")
                    proben = []; verworfen = 0
                continue

        proben.append(lab)
        print(f"          {len(proben)}/{WIEDERHOLUNGEN}   "
              f"L={lab[0]:6.2f}  a={lab[1]:6.2f}  b={lab[2]:6.2f}")

    m = mittel(proben)
    streuung = max(dE(p, m) for p in proben)
    abweichung = dE(m, erwartet)

    hinweis = ''
    treffer, td = bester_treffer(m)
    if treffer != ral and td < abweichung - 2:
        hinweis = f"passt besser zu RAL {treffer}"
    elif abweichung > DE_WARNUNG:
        hinweis = "grosse Abweichung zur Erwartung"

    print(f"          fertig:  L={m[0]:6.2f}  a={m[1]:6.2f}  b={m[2]:6.2f}"
          f"   Streuung {streuung:.2f}   Abweichung {abweichung:.1f}"
          + (f"   [{hinweis}]" if hinweis else ""))

    return m, streuung, abweichung, hinweis

# ---------------------------------------------------------------
# Excel-Mappe schreiben
# ---------------------------------------------------------------

def excel_schreiben(fertig):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\nopenpyxl fehlt - Excel uebersprungen.")
        print("Nachtraeglich:   python -m pip install openpyxl")
        print(f"Die Messwerte stehen vollstaendig in {FORTSCHRITT}.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Messwerte"

    kopf = ["RAL", "Farbname", "Soll L*", "Soll a*", "Soll b*",
            "Ist L*", "Ist a*", "Ist b*",
            "Streuung", "Abweichung dE", "Ist-Farbe", "Soll-Farbe", "Hinweis"]

    arial      = Font(name="Arial", size=10)
    arial_fett = Font(name="Arial", size=10, bold=True)
    kopf_fill  = PatternFill("solid", fgColor="D9D9D9")
    rahmen     = Border(*[Side(style="thin", color="BFBFBF")]*4)

    for c, t in enumerate(kopf, 1):
        z = ws.cell(row=1, column=c, value=t)
        z.font = arial_fett; z.fill = kopf_fill
        z.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        z.border = rahmen

    zeile = 2
    for ral, sL, sa, sb, shex, name in RAL_KARTE:
        d = fertig.get(ral)
        ws.cell(row=zeile, column=1, value=ral).font = arial
        ws.cell(row=zeile, column=2, value=name).font = arial
        for i, v in enumerate((sL, sa, sb), start=3):
            z = ws.cell(row=zeile, column=i, value=round(v, 2))
            z.font = arial; z.number_format = "0.00"

        if d:
            for i, k in enumerate(('L', 'a', 'b'), start=6):
                z = ws.cell(row=zeile, column=i, value=round(d[k], 2))
                z.font = arial; z.number_format = "0.00"
            z = ws.cell(row=zeile, column=9, value=round(d['streuung'], 2))
            z.font = arial; z.number_format = "0.00"
            # Abweichung als Formel, damit sie beim Aendern mitrechnet
            r = zeile
            z = ws.cell(row=r, column=10,
                        value=f"=IFERROR(SQRT((F{r}-C{r})^2+(G{r}-D{r})^2+(H{r}-E{r})^2),\"\")")
            z.font = arial; z.number_format = "0.0"
            ist_hex = lab_zu_hex(d['L'], d['a'], d['b'])
            zi = ws.cell(row=zeile, column=11)
            zi.fill = PatternFill("solid", fgColor=ist_hex)
            zi.border = rahmen
            if d['hinweis']:
                z = ws.cell(row=zeile, column=13, value=d['hinweis'])
                z.font = Font(name="Arial", size=10, color="C00000")
        else:
            z = ws.cell(row=zeile, column=13, value="noch nicht gemessen")
            z.font = Font(name="Arial", size=10, color="808080")

        zs = ws.cell(row=zeile, column=12)
        zs.fill = PatternFill("solid", fgColor=shex)
        zs.border = rahmen
        zeile += 1

    breiten = [8, 22, 9, 9, 9, 9, 9, 9, 10, 13, 11, 11, 26]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.freeze_panes = "A2"

    # --- Uebersichtsblatt ---
    ws2 = wb.create_sheet("Uebersicht")
    letzte = len(RAL_KARTE) + 1
    zeilen = [
        ("Farben gesamt",            f"=COUNT(Messwerte!A2:A{letzte})"),
        ("davon gemessen",           f"=COUNT(Messwerte!F2:F{letzte})"),
        ("noch offen",               f"=B2-B3"),
        ("", ""),
        ("Abweichung Mittel",        f"=IFERROR(AVERAGE(Messwerte!J2:J{letzte}),\"\")"),
        ("Abweichung groesster",     f"=IFERROR(MAX(Messwerte!J2:J{letzte}),\"\")"),
        ("Streuung Mittel",          f"=IFERROR(AVERAGE(Messwerte!I2:I{letzte}),\"\")"),
        ("Streuung groesster",       f"=IFERROR(MAX(Messwerte!I2:I{letzte}),\"\")"),
        ("", ""),
        ("Farben mit Hinweis",       f"=COUNTIF(Messwerte!M2:M{letzte},\"passt besser*\")"),
    ]
    ws2.cell(row=1, column=1, value="Auswertung").font = Font(name="Arial", size=12, bold=True)
    for i, (bez, formel) in enumerate(zeilen, start=2):
        ws2.cell(row=i, column=1, value=bez).font = arial
        if formel:
            z = ws2.cell(row=i, column=2, value=formel)
            z.font = arial; z.number_format = "0.00"
    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 14

    ws2.cell(row=14, column=1, value="Hinweise zur Verwendung").font = arial_fett
    for i, t in enumerate([
        "Soll-Werte stammen aus der oeffentlichen RAL-Tabelle (Naeherung aus Farbcodes).",
        "Ist-Werte sind die gemessenen Mittelwerte aus je drei Messungen.",
        "Streuung: groesster Abstand einer Einzelmessung zum Mittel. Unter 1,0 ist gut.",
        "Abweichung: Abstand des Messwerts zum Soll-Wert der Tabelle.",
        "Rot markierte Hinweise pruefen - dort passt der Messwert besser zu einer anderen Farbe.",
    ], start=15):
        ws2.cell(row=i, column=1, value="- " + t).font = arial

    wb.save(EXCEL)
    print(f"\nExcel-Mappe geschrieben: {EXCEL}")

# ---------------------------------------------------------------
# Hauptprogramm
# ---------------------------------------------------------------

def argumente():
    """Sehr einfache Auswertung: Port, --original DATEI, --nur-bauen."""
    port, original, nur_bauen = "COM3", None, False
    rest = sys.argv[1:]
    i = 0
    while i < len(rest):
        a = rest[i]
        if a == "--original" and i + 1 < len(rest):
            original = rest[i + 1]; i += 2
        elif a == "--nur-bauen":
            nur_bauen = True; i += 1
        elif not a.startswith("-"):
            port = a; i += 1
        else:
            sys.exit(f"Unbekannte Angabe: {a}")
    return port, original, nur_bauen


def firmware_bauen_wenn_moeglich(original, fertig):
    """Baut aus Originalfirmware und Messwerten die fertige Datei."""
    if not original:
        print("\nWenn du beim naechsten Mal --original DEINE_SICHERUNG.bin mit")
        print("angibst, baue ich die fertige Firmware gleich mit.")
        return
    if not os.path.exists(original):
        print(f"\nDatei nicht gefunden: {original}")
        return
    if len(fertig) < len(RAL_KARTE):
        print(f"\nErst {len(fertig)} von {len(RAL_KARTE)} Farben eingemessen.")
        print("Die fehlenden werden aus der mitgelieferten Vorlage uebernommen.")
    try:
        import firmware_bauen
    except ImportError:
        print("\nfirmware_bauen.py liegt nicht daneben - Firmware nicht gebaut.")
        return
    print("\n" + "=" * 62)
    print("  Firmware bauen")
    print("=" * 62)
    try:
        firmware_bauen.aus_dateien(original, FORTSCHRITT, "cr4501_ral.bin")
    except Exception as e:
        print(f"\nAbgebrochen: {e}")


def main():
    port, original, nur_bauen = argumente()

    fertig = fortschritt_laden()

    if nur_bauen:
        if not fertig:
            sys.exit(f"Keine Messwerte in {FORTSCHRITT}.")
        print(f"{len(fertig)} eingemessene Farben gefunden.")
        firmware_bauen_wenn_moeglich(original, fertig)
        return

    offen = [e for e in RAL_KARTE if e[0] not in fertig]

    print("=" * 62)
    print("  RAL-Faecher einlernen  -  3nh CR4501")
    print("=" * 62)
    if fertig:
        print(f"  Fortsetzung: {len(fertig)} Farben bereits erfasst, "
              f"{len(offen)} offen.")
    else:
        print(f"  {len(RAL_KARTE)} Farben, je {WIEDERHOLUNGEN} Messungen "
              f"= {len(RAL_KARTE)*WIEDERHOLUNGEN} Messungen.")
    print(f"  Reihenfolge wie auf der Karte: RAL {RAL_KARTE[0][0]} "
          f"bis {RAL_KARTE[-1][0]}")
    print("  Abbruch mit Strg+C - der Stand bleibt erhalten.")
    print("=" * 62)

    if not offen:
        print("\nAlle Farben sind bereits erfasst.")
        excel_schreiben(fertig)
        firmware_bauen_wenn_moeglich(original, fertig)
        return

    try:
        dev = Messgeraet(port)
    except Exception as e:
        sys.exit(f"\nKonnte {port} nicht oeffnen: {e}\n"
                 f"Richtigen Anschluss im Geraete-Manager nachsehen.")

    dev.leeren()
    vorfarbe = None
    start = time.time()

    try:
        for i, eintrag in enumerate(offen, start=1):
            ergebnis = farbe_einlernen(dev, eintrag, vorfarbe, i, len(offen))
            if ergebnis is None:
                continue
            m, streuung, abweichung, hinweis = ergebnis

            fertig[eintrag[0]] = {'L': m[0], 'a': m[1], 'b': m[2],
                                  'streuung': streuung,
                                  'abweichung': abweichung,
                                  'hinweis': hinweis}
            fortschritt_schreiben(fertig)     # nach jeder Farbe sichern
            vorfarbe = m

            if i % 10 == 0:
                pro = (time.time() - start) / i
                rest = (len(offen) - i) * pro / 60
                print(f"\n          --- {i} von {len(offen)} erledigt, "
                      f"noch etwa {rest:.0f} Minuten ---")

    except KeyboardInterrupt:
        print("\n\nAbgebrochen. Stand gesichert.")
        print(f"Weiter mit demselben Befehl - es geht bei RAL "
              f"{offen[len(fertig) - (len(RAL_KARTE) - len(offen))][0] if offen else '?'} weiter.")
    finally:
        dev.schliessen()

    print(f"\n{len(fertig)} von {len(RAL_KARTE)} Farben erfasst.")
    excel_schreiben(fertig)

    firmware_bauen_wenn_moeglich(original, fertig)


if __name__ == "__main__":
    main()
